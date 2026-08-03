import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


# ==========================================================
# MAPEAMENTO DAS COLUNAS
# ==========================================================

COLUMNS_MAP = {
    "uc": "UC",
    "cliente": "Cliente",
    "medidor": "Medidor",
    "local_unidade": "Local",
    "conta_mes": "Referência",
    "vencimento": "Vencimento",
    "classificacao": "Classificação",
    "tipo_fornecimento": "Tipo de Fornecimento",
    "leitura_anterior": "Leitura Anterior",
    "leitura_atual": "Leitura Atual",
    "consumo_kwh": "Consumo (kWh)",
    "total_pagar": "Total a Pagar",
}


# ==========================================================
# TIPOS DE COLUNAS
# ==========================================================

TEXT_COLUMNS = [
    "Cliente",
    "UC",
    "Medidor",
    "Local",
    "Referência",
    "Vencimento",
    "Classificação",
    "Tipo de Fornecimento",
]

NUMERIC_COLUMNS = [
    "Leitura Anterior",
    "Leitura Atual",
    "Consumo (kWh)",
]

CURRENCY_COLUMNS = [
    "Total a Pagar",
]


MUNICIPALITY_PREFIX = re.compile(r"^\s*MUNIC(?:I|Í)PIO\s*DE\s*", re.IGNORECASE)


def _normalize_export_cliente(value: str) -> str:
    """Remove the municipality prefix only in the Excel export copy."""
    return MUNICIPALITY_PREFIX.sub("", value).strip()


# ==========================================================
# ESCREVER NO EXCEL
# ==========================================================

def write_to_excel(data_list: list, output_path: str) -> str:
    """
    Grava ou atualiza os dados extraídos das faturas em um Excel.

    Duplicidades são identificadas pela combinação:
    UC + Referência

    Se a mesma UC e Referência já existirem,
    o registro mais recente substitui o antigo.
    """

    logger.info(
        f"Gravando {len(data_list)} registros no Excel: "
        f"{output_path}"
    )

    # ======================================================
    # VALIDAR DADOS
    # ======================================================

    if not data_list:

        logger.warning(
            "Nenhum dado recebido para gravar no Excel."
        )

        return os.path.abspath(output_path)

    # ======================================================
    # CRIAR DATAFRAME
    # ======================================================

    df_new_raw = pd.DataFrame(data_list)

    # Garante que todas as colunas esperadas existam
    for column in COLUMNS_MAP.keys():

        if column not in df_new_raw.columns:

            df_new_raw[column] = "N/A"

    # Renomeia para os nomes amigáveis
    df_new = df_new_raw.rename(
        columns=COLUMNS_MAP
    )

    # Mantém somente as colunas configuradas
    df_new = df_new[
        list(COLUMNS_MAP.values())
    ]

    # ======================================================
    # NORMALIZAR NOVOS DADOS
    # ======================================================

    for column in df_new.columns:

        df_new[column] = (
            df_new[column]
            .fillna("N/A")
            .astype(str)
            .str.strip()
        )

    # This DataFrame is an export-only snapshot.  Do not modify the source
    # objects received from the editable table in the browser.
    df_new["Cliente"] = df_new["Cliente"].map(_normalize_export_cliente)

    # Export is a snapshot of the table sent by the UI.  UC and reference
    # month are not unique in grouped PDFs, so preserve every supplied row.
    df_final = df_new

    # ======================================================
    # CONVERTER NÚMEROS
    # ======================================================

    for column in NUMERIC_COLUMNS:

        df_final[column] = pd.to_numeric(
            df_final[column],
            errors="coerce"
        ).fillna(0)

    for column in CURRENCY_COLUMNS:

        df_final[column] = pd.to_numeric(
            df_final[column],
            errors="coerce"
        ).fillna(0.0)

    # ======================================================
    # SALVAR EXCEL
    # ======================================================

    df_final.to_excel(
        output_path,
        index=False
    )

    # ======================================================
    # APLICAR ESTILO
    # ======================================================

    apply_premium_styling(
        output_path
    )

    logger.info(
        "Planilha Excel salva com sucesso."
    )

    return os.path.abspath(
        output_path
    )


# ==========================================================
# ESTILIZAÇÃO
# ==========================================================

def apply_premium_styling(file_path: str):

    """
    Aplica estilo visual à planilha Excel.
    """

    wb = openpyxl.load_workbook(
        file_path
    )

    ws = wb.active

    # Nome da aba
    ws.title = "Faturas Consolidadas"

    # ======================================================
    # ESTILOS
    # ======================================================

    font_family = "Segoe UI"

    header_font = Font(
        name=font_family,
        size=11,
        bold=True,
        color="FFFFFF"
    )

    data_font = Font(
        name=font_family,
        size=10,
        color="333333"
    )

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    zebra_fill = PatternFill(
        start_color="F2F5F8",
        end_color="F2F5F8",
        fill_type="solid"
    )

    thin_border = Border(
        left=Side(
            style="thin",
            color="D9D9D9"
        ),
        right=Side(
            style="thin",
            color="D9D9D9"
        ),
        top=Side(
            style="thin",
            color="D9D9D9"
        ),
        bottom=Side(
            style="thin",
            color="D9D9D9"
        )
    )

    # ======================================================
    # CABEÇALHO
    # ======================================================

    for col_idx in range(
        1,
        ws.max_column + 1
    ):

        cell = ws.cell(
            row=1,
            column=col_idx
        )

        cell.font = header_font

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    # ======================================================
    # LINHAS
    # ======================================================

    for row_idx in range(
        2,
        ws.max_row + 1
    ):

        ws.row_dimensions[
            row_idx
        ].height = 20

        is_zebra = (
            row_idx % 2 == 0
        )

        for col_idx in range(
            1,
            ws.max_column + 1
        ):

            cell = ws.cell(
                row=row_idx,
                column=col_idx
            )

            header_name = ws.cell(
                row=1,
                column=col_idx
            ).value

            cell.font = data_font

            cell.border = thin_border

            if is_zebra:

                cell.fill = zebra_fill

            # ------------------------------
            # TEXTO
            # ------------------------------

            if header_name in TEXT_COLUMNS:

                cell.number_format = "@"

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # ------------------------------
            # NÚMEROS
            # ------------------------------

            elif header_name in NUMERIC_COLUMNS:

                cell.number_format = "#,##0"

                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center"
                )

            # ------------------------------
            # MOEDA
            # ------------------------------

            elif header_name in CURRENCY_COLUMNS:

                cell.number_format = (
                    'R$ #,##0.00'
                )

                cell.alignment = Alignment(
                    horizontal="right",
                    vertical="center"
                )

            else:

                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )

    # ======================================================
    # AUTO AJUSTE DAS COLUNAS
    # ======================================================

    for col in ws.columns:

        max_len = 0

        col_letter = get_column_letter(
            col[0].column
        )

        for cell in col:

            value = cell.value

            if value is None:

                value = ""

            value_str = str(
                value
            )

            if (
                cell.number_format
                == 'R$ #,##0.00'
                and isinstance(
                    value,
                    (int, float)
                )
            ):

                value_str = (
                    f"R$ {value:,.2f}"
                    .replace(",", "X")
                    .replace(".", ",")
                    .replace("X", ".")
                )

            max_len = max(
                max_len,
                len(value_str)
            )

        ws.column_dimensions[
            col_letter
        ].width = max(
            max_len + 4,
            12
        )

    # ======================================================
    # CONGELAR CABEÇALHO
    # ======================================================

    ws.freeze_panes = "A2"

    # ======================================================
    # FILTRO
    # ======================================================

    ws.auto_filter.ref = (
        ws.dimensions
    )

    # ======================================================
    # SALVAR
    # ======================================================

    wb.save(
        file_path
    )
