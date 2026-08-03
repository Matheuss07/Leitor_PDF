from copy import deepcopy

import openpyxl

from app.core.excel_writer import write_to_excel


def test_excel_export_excludes_columns_and_normalizes_municipality_client(tmp_path):
    data = [{
        "filename": "fatura_01.pdf",
        "uc": "669.008.008-24",
        "cliente": "MUNICIPIO DE FLEXEIRAS",
        "medidor": "123456",
        "local_unidade": "R. CEL ALCANTARA, S/N",
        "conta_mes": "06/2026",
        "vencimento": "2026-07-23",
        "classificacao": "PODER PÚBLICO MUNICIPAL",
        "subclasse": "MUNICIPAL SERVICOS PUBLICOS",
        "tipo_fornecimento": "MONOFÁSICO",
        "leitura_anterior": 1000,
        "leitura_atual": 1200,
        "consumo_kwh": 200,
        "total_pagar": 350.0,
    }]
    original_data = deepcopy(data)
    output_path = tmp_path / "faturas.xlsx"

    write_to_excel(data, str(output_path))

    worksheet = openpyxl.load_workbook(output_path, data_only=True).active
    headers = [cell.value for cell in worksheet[1]]
    exported_row = [cell.value for cell in worksheet[2]]

    assert "Arquivo" not in headers
    assert "Subclasse" not in headers
    assert exported_row[headers.index("Cliente")] == "FLEXEIRAS"
    assert data == original_data


def test_excel_export_keeps_non_municipality_client_unchanged(tmp_path):
    output_path = tmp_path / "faturas.xlsx"
    data = [{"cliente": "PREFEITURA DE FLEXEIRAS"}]

    write_to_excel(data, str(output_path))

    worksheet = openpyxl.load_workbook(output_path, data_only=True).active
    headers = [cell.value for cell in worksheet[1]]
    assert worksheet.cell(2, headers.index("Cliente") + 1).value == "PREFEITURA DE FLEXEIRAS"


def test_excel_export_normalizes_municipality_prefix_without_spaces(tmp_path):
    output_path = tmp_path / "faturas.xlsx"
    data = [
        {"cliente": "MUNICIPIODEFLEXEIRAS"},
        {"cliente": "MUNICÍPIODEFLEXEIRAS"},
        {"cliente": "MUNICIPIO  DE  FLEXEIRAS"},
        {"cliente": "MUNICIPIO ABC"},
    ]

    write_to_excel(data, str(output_path))

    worksheet = openpyxl.load_workbook(output_path, data_only=True).active
    headers = [cell.value for cell in worksheet[1]]
    client_column = headers.index("Cliente") + 1
    exported_clients = [
        worksheet.cell(row, client_column).value
        for row in range(2, 6)
    ]
    assert exported_clients == [
        "FLEXEIRAS", "FLEXEIRAS", "FLEXEIRAS", "MUNICIPIO ABC",
    ]
